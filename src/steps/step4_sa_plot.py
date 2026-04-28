from __future__ import annotations

import os
from multiprocessing import Pool, cpu_count
from pathlib import Path
from typing import Any, Dict, Tuple

import numpy as np
from openpyxl import Workbook

from src.lib.psa import compute_psa


def _process_one_folder_job(args: Tuple[str, str, Dict[str, Any]]):
    folder_path_str, output_dir_str, cfg = args
    folder_path = Path(folder_path_str)
    output_dir = Path(output_dir_str)

    pid = os.getpid()

    gm_files = sorted(folder_path.glob("*.txt"))
    if not gm_files:
        return (folder_path.name, "SKIP(no_txt)")

    dt = float(cfg["dt"])
    zeta = float(cfg["zeta"])
    T_min = float(cfg["T_min"])
    T_max = float(cfg["T_max"])
    nT = int(cfg["nT"])
    T_array = np.logspace(np.log10(T_min), np.log10(T_max), nT)

    export_format = str(cfg.get("export", {}).get("format", "xlsx")).lower()

    if export_format == "npz":
        Sa_mat = np.zeros((len(gm_files), len(T_array)), dtype=np.float32)
        names = []
        for i, gm_path in enumerate(gm_files):
            acc = np.loadtxt(gm_path)
            Sa_mat[i, :] = compute_psa(acc, dt, T_array, zeta=zeta).astype(np.float32)
            names.append(gm_path.stem)
        out_path = output_dir / f"{folder_path.name}.npz"
        np.savez_compressed(out_path, T=T_array.astype(np.float32), Sa=Sa_mat, names=np.array(names))
        return (folder_path.name, "OK(npz)")

    # default: xlsx
    excel_path = output_dir / f"{folder_path.name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PSA"

    ws["A1"] = "T (sec)"
    for i, T in enumerate(T_array, start=2):
        ws[f"A{i}"] = float(T)

    col_idx = 2
    for gm_path in gm_files:
        acc = np.loadtxt(gm_path)
        Sa = compute_psa(acc, dt, T_array, zeta=zeta)

        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws[f"{col_letter}1"] = gm_path.stem
        for i, sa in enumerate(Sa, start=2):
            ws[f"{col_letter}{i}"] = float(sa)
        col_idx += 1

    wb.save(excel_path)
    if len(gm_files) > 0:
        print(f"[PID {pid}] Saved {excel_path.name} (files={len(gm_files)})")
    return (folder_path.name, "OK(xlsx)")


def export_psa(cfg: Dict[str, Any], root_input_dir: Path, output_dir: Path, nproc: int | None = None) -> Dict[str, Any]:
    if not root_input_dir.exists():
        raise FileNotFoundError(f"No input dir: {root_input_dir}")

    subfolders = sorted([p for p in root_input_dir.iterdir() if p.is_dir()])
    if not subfolders:
        return {"num_folders": 0, "status": []}

    output_dir.mkdir(parents=True, exist_ok=True)

    if nproc is None:
        nproc = min(cpu_count(), len(subfolders))
    else:
        nproc = max(1, min(int(nproc), len(subfolders)))

    jobs = [(str(folder), str(output_dir), cfg) for folder in subfolders]

    with Pool(processes=nproc) as pool:
        results = pool.map(_process_one_folder_job, jobs)

    return {"num_folders": len(subfolders), "results": results, "out_dir": str(output_dir)}
